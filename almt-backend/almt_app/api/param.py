"""
参数配置API
"""
from fastapi import APIRouter
from typing import List
import pymysql
import uuid

router = APIRouter(prefix="/api/param", tags=["参数配置"])


def get_db_conn():
    return pymysql.connect(
        host='localhost', user='almt', password='almt',
        database='almt_db', port=3306, cursorclass=pymysql.cursors.DictCursor
    )


# ============ 利率情景假设 ============
M_FIELDS = ['m' + str(i) + '_value' for i in range(1, 25)]

@router.get("/rate-scenario")
def list_rate_scenario(skip: int = 0, limit: int = 200, search: str = ''):
    """获取利率情景列表"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            if search:
                cursor.execute(
                    """SELECT * FROM almt_param_rate_scenario
                    WHERE curve_name LIKE %s OR curve_id LIKE %s OR scenario_name LIKE %s
                    ORDER BY scenario_name, order_number LIMIT %s OFFSET %s""",
                    ('%' + search + '%', '%' + search + '%', '%' + search + '%', limit, skip)
                )
            else:
                cursor.execute(
                    """SELECT * FROM almt_param_rate_scenario
                    ORDER BY scenario_name, order_number LIMIT %s OFFSET %s""",
                    (limit, skip)
                )
            return cursor.fetchall()
    finally:
        conn.close()


@router.post("/rate-scenario")
def create_rate_scenario(item: dict):
    """创建利率情景（自动根据 current_value + shift 计算 M1-M24）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            uuid_val = item.get('uuid') or str(uuid.uuid4())
            current_val = float(item.get('current_curve_value', 0))
            shift = float(item.get('scenario_shift', 0))
            m_values = [current_val + shift for _ in range(24)]
            # 如果客户端传了m_values则优先使用
            for i, f in enumerate(M_FIELDS):
                if item.get(f) is not None:
                    m_values[i] = float(item[f])

            fields = ['uuid', 'order_number', 'curve_name', 'curve_id', 'scenario_name', 'scenario_shift', 'current_curve_value', 'remark'] + M_FIELDS
            values = [uuid_val, item.get('order_number'), item.get('curve_name', ''),
                      item.get('curve_id', ''), item.get('scenario_name', '基准'),
                      shift, current_val, item.get('remark', '')] + m_values
            placeholders = ','.join(['%s'] * len(fields))
            sql = 'INSERT INTO almt_param_rate_scenario (' + ','.join(fields) + ') VALUES (' + placeholders + ')'
            cursor.execute(sql, values)
        conn.commit()
        return {"message": "创建成功", "uuid": uuid_val}
    finally:
        conn.close()


@router.put("/rate-scenario/{item_id}")
def update_rate_scenario(item_id: int, item: dict):
    """更新利率情景（自动根据 current_value + shift 计算 M1-M24）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            current_val = float(item.get('current_curve_value', 0))
            shift = float(item.get('scenario_shift', 0))
            m_values = [current_val + shift for _ in range(24)]
            for i, f in enumerate(M_FIELDS):
                if item.get(f) is not None:
                    m_values[i] = float(item[f])

            fields = ['order_number', 'curve_name', 'curve_id', 'scenario_name', 'scenario_shift', 'current_curve_value', 'remark'] + M_FIELDS
            values = [item.get('order_number'), item.get('curve_name', ''),
                      item.get('curve_id', ''), item.get('scenario_name', '基准'),
                      shift, current_val, item.get('remark', '')] + m_values
            values.append(item_id)
            set_clause = ','.join([f + '=%s' for f in fields])
            sql = 'UPDATE almt_param_rate_scenario SET ' + set_clause + ' WHERE id=%s'
            cursor.execute(sql, values)
        conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@router.delete("/rate-scenario/{item_id}")
def delete_rate_scenario(item_id: int):
    """删除利率情景"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM almt_param_rate_scenario WHERE id=%s", (item_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


# ============ 风险权重 ============
@router.get("/risk-weight")
def list_risk_weight():
    """获取风险权重列表"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM almt_param_risk_weight LIMIT 100")
            return cursor.fetchall()
    finally:
        conn.close()


@router.post("/risk-weight")
def create_risk_weight(item: dict):
    """创建风险权重"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            uuid_val = item.get('uuid') or str(uuid.uuid4())
            cursor.execute(
                """INSERT INTO almt_param_risk_weight (uuid, coa_cd, coa_name, weight) 
                VALUES (%s, %s, %s, %s)""",
                (uuid_val, item.get('coa_cd', ''), item.get('coa_name', ''), item.get('weight', 0))
            )
        conn.commit()
        return {"message": "创建成功", "uuid": uuid_val}
    finally:
        conn.close()


@router.put("/risk-weight/{item_id}")
def update_risk_weight(item_id: int, item: dict):
    """更新风险权重"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """UPDATE almt_param_risk_weight SET coa_cd=%s, coa_name=%s, weight=%s WHERE id=%s""",
                (item.get('coa_cd', ''), item.get('coa_name', ''), item.get('weight', 0), item_id)
            )
        conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@router.post("/risk-weight/save")
def save_risk_weight(item: dict):
    """保存风险权重（按 coa_cd upsert），含24期 risk_weight_1~24"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            coa_cd = item.get('coa_cd', '')
            coa_name = item.get('coa_name', '')
            cols = [item.get(f'risk_weight_{i}') for i in range(1, 25)]
            set_clause = ', '.join([f'risk_weight_{i}=%s' for i in range(1, 25)])
            cursor.execute("SELECT id FROM almt_param_risk_weight WHERE coa_cd=%s", (coa_cd,))
            existing = cursor.fetchone()
            if existing:
                # 同步 weight 字段为 M1
                cursor.execute(
                    f"""UPDATE almt_param_risk_weight SET coa_name=%s, weight=%s, {set_clause} WHERE id=%s""",
                    (coa_name, cols[0] or 0) + tuple(cols) + (existing['id'],)
                )
                action = '更新'
            else:
                placeholders = ', '.join(['%s'] * 24)
                risk_cols = ', '.join([f'risk_weight_{i}' for i in range(1, 25)])
                cursor.execute(
                    f"""INSERT INTO almt_param_risk_weight (uuid, coa_cd, coa_name, weight, {risk_cols})
                        VALUES (%s, %s, %s, %s, {placeholders})""",
                    (str(uuid.uuid4()), coa_cd, coa_name, cols[0] or 0) + tuple(cols)
                )
                action = '创建'
        conn.commit()
        return {"message": f"{action}成功"}
    finally:
        conn.close()


@router.delete("/risk-weight/{item_id}")
def delete_risk_weight(item_id: int):
    """删除风险权重"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM almt_param_risk_weight WHERE id=%s", (item_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


# ============ 业务计划 ============
@router.get("/business-plan")
def list_business_plan():
    """获取业务计划列表"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM almt_param_business_plan ORDER BY id LIMIT 100")
            return cursor.fetchall()
    finally:
        conn.close()


@router.post("/business-plan")
def create_business_plan(item: dict):
    """创建业务计划"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            uuid_val = item.get('uuid') or str(uuid.uuid4())
            cursor.execute(
                """INSERT INTO almt_param_business_plan (uuid, coa_lvl, coa_cd, coa_name) 
                VALUES (%s, %s, %s, %s)""",
                (uuid_val, item.get('coa_lvl', ''), item.get('coa_cd', ''), item.get('coa_name', ''))
            )
        conn.commit()
        return {"message": "创建成功", "uuid": uuid_val}
    finally:
        conn.close()


@router.post("/business-plan/save")
def save_business_plan(item: dict):
    """保存业务计划（按 coa_cd upsert），含24期日均增量"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            coa_cd = item.get('coa_cd', '')
            coa_name = item.get('coa_name', '')
            coa_lvl = item.get('coa_lvl', '')
            # 收集24期日均增量字段
            avg_cols = [item.get(f'average_{i}', None) for i in range(1, 25)]
            bal_cols = [item.get(f'balance_{i}', None) for i in range(1, 25)]

            # 构建 update/insert SQL
            set_avg = ', '.join([f'plan_average{i}=%s' for i in range(1, 25)])
            set_bal = ', '.join([f'plan_balance{i}=%s' for i in range(1, 25)])
            cursor.execute("SELECT id FROM almt_param_business_plan WHERE coa_cd=%s", (coa_cd,))
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    f"""UPDATE almt_param_business_plan
                        SET coa_name=%s, coa_lvl=%s, {set_avg}, {set_bal}
                        WHERE id=%s""",
                    (coa_name, coa_lvl) + tuple(avg_cols) + tuple(bal_cols) + (existing['id'],)
                )
                action = '更新'
            else:
                avg_placeholders = ', '.join(['%s'] * 24)
                bal_placeholders = ', '.join(['%s'] * 24)
                avg_cols_full = [f'plan_average{i}' for i in range(1, 25)]
                bal_cols_full = [f'plan_balance{i}' for i in range(1, 25)]
                cursor.execute(
                    f"""INSERT INTO almt_param_business_plan
                        (uuid, coa_lvl, coa_cd, coa_name,
                         {', '.join(avg_cols_full)},
                         {', '.join(bal_cols_full)})
                        VALUES (%s, %s, %s, %s, {avg_placeholders}, {bal_placeholders})""",
                    (str(uuid.uuid4()), coa_lvl, coa_cd, coa_name) + tuple(avg_cols) + tuple(bal_cols)
                )
                action = '创建'
        conn.commit()
        return {"message": f"{action}成功"}
    finally:
        conn.close()


# ============ FTP利差 ============
@router.get("/ftp-margin")
def list_ftp_margin():
    """获取FTP利差列表"""
    return []


@router.post("/ftp-margin")
def create_ftp_margin(item: dict):
    """创建FTP利差"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """INSERT INTO almt_param_ftp_spread (uuid, coa_cd, coa_name, spread, curve_name, term)
                VALUES (%s, %s, %s, %s, %s, %s)""",
                (str(uuid.uuid4()), item.get('coa_cd', ''), item.get('coa_name', ''),
                 item.get('spread', 0), item.get('curve_name', ''), item.get('term', ''))
            )
        conn.commit()
        return {"message": "创建成功"}
    finally:
        conn.close()


@router.post("/ftp-margin/save")
def save_ftp_margin(item: dict):
    """保存FTP利差（按 coa_cd upsert）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            coa_cd = item.get('coa_cd', '')
            spread = item.get('spread', 0)
            coa_name = item.get('coa_name', '')
            curve_name = item.get('curve_name', '')
            term = item.get('term', '')
            cursor.execute("SELECT id FROM almt_param_ftp_spread WHERE coa_cd=%s", (coa_cd,))
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    """UPDATE almt_param_ftp_spread SET coa_name=%s, spread=%s, curve_name=%s, term=%s WHERE id=%s""",
                    (coa_name, spread, curve_name, term, existing['id'])
                )
                action = '更新'
            else:
                cursor.execute(
                    """INSERT INTO almt_param_ftp_spread (uuid, coa_cd, coa_name, spread, curve_name, term) 
                    VALUES (%s, %s, %s, %s, %s, %s)""",
                    (str(uuid.uuid4()), coa_cd, coa_name, spread, curve_name, term)
                )
                action = '创建'
        conn.commit()
        return {"message": f"{action}成功"}
    finally:
        conn.close()


@router.put("/ftp-margin/{item_id}")
def update_ftp_margin(item_id: int, item: dict):
    """更新FTP利差"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """UPDATE almt_param_ftp_spread SET coa_cd=%s, coa_name=%s, spread=%s, curve_name=%s, term=%s WHERE id=%s""",
                (item.get('coa_cd', ''), item.get('coa_name', ''),
                 item.get('spread', 0), item.get('curve_name', ''), item.get('term', ''), item_id)
            )
        conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@router.delete("/ftp-margin/{item_id}")
def delete_ftp_margin(item_id: int):
    """删除FTP利差"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM almt_param_ftp_spread WHERE id=%s", (item_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


# ============ 树形结构接口（按账户册层级展示）============
def build_tree_from_coa(all_coa: list, value_map: dict, value_field: str = 'value') -> list:
    """根据账户册和数值映射构建树形结构"""
    coa_map = {}
    roots = []

    for coa in all_coa:
        node = {
            "id": coa['id'],
            "coa_cd": coa['coa_cd'],
            "coa_name": coa['coa_name'],
            "leaf_flag": coa['leaf_flag'],
            "value": None,
            "has_data": False,
            "children": []
        }
        coa_map[coa['coa_cd']] = node

        if coa['coa_cd'] in value_map:
            v = value_map[coa['coa_cd']]
            node['value'] = float(v) if v is not None else None
            node['has_data'] = True

    # 用parent_coa_cd构建父子关系
    for coa in all_coa:
        node = coa_map[coa['coa_cd']]
        parent_cd = coa['parent_coa_cd']
        if not parent_cd or parent_cd not in coa_map:
            roots.append(node)
        else:
            coa_map[parent_cd]['children'].append(node)

    # 自底向上汇总（如果是数值类型）
    if value_field == 'value':
        def calc_sum(node):
            for child in node['children']:
                calc_sum(child)
            if not node['children']:
                return
            total = 0
            any_data = False
            for child in node['children']:
                if child['value'] is not None:
                    total += child['value']
                    any_data = True
            if any_data:
                node['value'] = total
                node['has_data'] = True
        for root in roots:
            calc_sum(root)

    return roots


@router.get("/risk-weight/tree")
def get_risk_weight_tree():
    """按账户册树形展示风险权重（24期，不汇总，支持父级继承）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, coa_cd, coa_name, leaf_flag, parent_coa_cd FROM almt_coa_info")
            all_coa = cursor.fetchall()
            cols = ', '.join([f'risk_weight_{i}' for i in range(1, 25)])
            cursor.execute(f"SELECT coa_cd, {cols} FROM almt_param_risk_weight")
            weight_map = {}
            for row in cursor.fetchall():
                weight_map[row['coa_cd']] = [row[f'risk_weight_{i}'] for i in range(1, 25)]

        # 不进行汇总（保持原始 null/0 值），构建树
        # value_map 仅用于显示 has_data 状态，用 M1 的 risk_weight_1 作为标记
        value_map = {}
        for cd, ws in weight_map.items():
            value_map[cd] = ws[0] if (ws and ws[0] is not None) else None
        tree = build_tree_from_coa(all_coa, value_map)

        # 构建 parent 索引
        coa_index = {c['coa_cd']: c for c in all_coa}

        # 计算每个节点的 effective_weights（自身 OR 最近的祖先）和 inherited 标记
        def find_inherited_weights(cd: str):
            """从该节点向上找，直到找到任何一个 risk_weight_i != None 的祖先"""
            cur = cd
            while cur:
                node = coa_index.get(cur)
                if not node:
                    return [None] * 24, False
                # 如果当前节点有值（任何一个 risk_weight_i 不为 None）
                if cur in weight_map:
                    ws = weight_map[cur]
                    if any(w is not None for w in ws):
                        return ws, (cur == cd)
                # 向上递归
                parent_cd = node.get('parent_coa_cd')
                cur = parent_cd
            return [None] * 24, False

        # 为每个节点附加 24 期权重
        def attach_weights(nodes):
            for n in nodes:
                ws, is_self = find_inherited_weights(n['coa_cd'])
                n['weights'] = ws
                n['is_self'] = is_self
                if n.get('children'):
                    attach_weights(n['children'])
        attach_weights(tree)
        return tree
    finally:
        conn.close()


@router.get("/risk-weight/{coa_cd}")
def get_risk_weight_detail(coa_cd: str):
    """获取某个账户册的24期风险权重详情"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cols = ', '.join([f'risk_weight_{i}' for i in range(1, 25)])
            cursor.execute(
                f"SELECT coa_cd, coa_name, {cols} FROM almt_param_risk_weight WHERE coa_cd=%s",
                (coa_cd,)
            )
            row = cursor.fetchone()
            if not row:
                return {'coa_cd': coa_cd, 'coa_name': '', 'weights': [None]*24}
            return {
                'coa_cd': row['coa_cd'],
                'coa_name': row['coa_name'] or '',
                'weights': [row[f'risk_weight_{i}'] for i in range(1, 25)]
            }
    finally:
        conn.close()


@router.get("/ftp-margin/tree")
def get_ftp_margin_tree():
    """按账户册树形展示FTP利差"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, coa_cd, coa_name, leaf_flag, parent_coa_cd FROM almt_coa_info")
            all_coa = cursor.fetchall()
            cursor.execute("SELECT coa_cd, spread FROM almt_param_ftp_spread")
            value_map = {row['coa_cd']: row['spread'] for row in cursor.fetchall()}
        return build_tree_from_coa(all_coa, value_map)
    finally:
        conn.close()


@router.get("/business-plan/tree")
def get_business_plan_tree():
    """按账户册树形展示业务计划（含24期规模增量+日均增量明细）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, coa_cd, coa_name, leaf_flag, parent_coa_cd FROM almt_coa_info")
            all_coa = cursor.fetchall()
            # 读24期规模增量 + 24期日均增量
            bal_cols = ', '.join([f'plan_balance{i}' for i in range(1, 25)])
            avg_cols = ', '.join([f'plan_average{i}' for i in range(1, 25)])
            cursor.execute(
                f"SELECT coa_cd, coa_name, {bal_cols}, {avg_cols} FROM almt_param_business_plan"
            )
            plan_map = {}
            for row in cursor.fetchall():
                plan_map[row['coa_cd']] = {
                    'balances': [float(row[f'plan_balance{i}'] or 0) for i in range(1, 25)],
                    'averages': [float(row[f'plan_average{i}'] or 0) for i in range(1, 25)]
                }
        # value_map 用 M1 任意一个非零值作代表（优先规模增量）
        value_map = {}
        for k, v in plan_map.items():
            m1 = v['balances'][0] if v['balances'] else 0
            if not m1:
                m1 = v['averages'][0] if v['averages'] else 0
            value_map[k] = m1
        tree = build_tree_from_coa(all_coa, value_map)
        # 在每个有数据的节点上额外附加 24 期规模增量+日均增量
        def attach_plans(nodes):
            for n in nodes:
                if n.get('coa_cd') in plan_map:
                    n['balances'] = plan_map[n['coa_cd']]['balances']
                    n['averages'] = plan_map[n['coa_cd']]['averages']
                if n.get('children'):
                    attach_plans(n['children'])
        attach_plans(tree)
        return tree
    finally:
        conn.close()


@router.get("/business-plan/{coa_cd}")
def get_business_plan_detail(coa_cd: str):
    """获取某个账户册的业务计划24期明细"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cols = ', '.join([f'plan_balance{i}, plan_average{i}' for i in range(1, 25)]).replace(', ', ', ')
            cols = ', '.join([f'plan_balance{i}' for i in range(1, 25)] + [f'plan_average{i}' for i in range(1, 25)])
            cursor.execute(
                f"SELECT coa_lvl, coa_cd, coa_name, {cols} FROM almt_param_business_plan WHERE coa_cd=%s",
                (coa_cd,)
            )
            row = cursor.fetchone()
            if not row:
                return {'coa_lvl': '', 'coa_cd': coa_cd, 'coa_name': '', 'balances': [None]*24, 'averages': [None]*24}
            return {
                'coa_lvl': row['coa_lvl'] or '',
                'coa_cd': row['coa_cd'],
                'coa_name': row['coa_name'] or '',
                'balances': [row[f'plan_balance{i}'] for i in range(1, 25)],
                'averages': [row[f'plan_average{i}'] for i in range(1, 25)]
            }
    finally:
        conn.close()

# ============ 通用导入导出 ============
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
import io


def _export_excel(rows, headers, sheet_name, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/risk-weight/export")
def export_risk_weight():
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cols = ', '.join([f'risk_weight_{i}' for i in range(1, 25)])
            cursor.execute(f"SELECT coa_cd, coa_name, {cols} FROM almt_param_risk_weight ORDER BY coa_cd")
            raw = cursor.fetchall()
            headers = ['账户册编码', '账户册名称'] + [f'M{i}风险权重' for i in range(1, 25)]
            rows = []
            for r in raw:
                row = [r['coa_cd'], r['coa_name'] or '']
                for i in range(1, 25):
                    v = r.get(f'risk_weight_{i}')
                    row.append(float(v) if v is not None else None)
                rows.append(row)
    finally:
        conn.close()
    return _export_excel(rows, headers, "风险权重-24期", "risk_weight_24.xlsx")


@router.post("/risk-weight/import")
async def import_risk_weight(file):
    """导入风险权重：前2列 编码/名称，后续24列 M1~M24 risk_weight"""
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    conn = get_db_conn()
    try:
        ins, upd = 0, 0
        with conn.cursor() as cursor:
            for row in rows:
                if not row or not row[0]:
                    continue
                vals = list(row) + [None] * (26 - len(row))
                coa_cd = str(vals[0])
                coa_name = str(vals[1] or '')
                ws = []
                for i in range(2, 26):
                    try:
                        ws.append(float(vals[i]) if vals[i] not in (None, '') else None)
                    except Exception:
                        ws.append(None)
                m1 = ws[0] if ws and ws[0] is not None else 0
                set_clause = ', '.join([f'risk_weight_{i}=%s' for i in range(1, 25)])
                cursor.execute("SELECT id FROM almt_param_risk_weight WHERE coa_cd=%s", (coa_cd,))
                ex = cursor.fetchone()
                if ex:
                    cursor.execute(
                        f"UPDATE almt_param_risk_weight SET coa_name=%s, weight=%s, {set_clause} WHERE id=%s",
                        (coa_name, m1) + tuple(ws) + (ex['id'],)
                    )
                    upd += 1
                else:
                    placeholders = ', '.join(['%s'] * 24)
                    risk_cols = ', '.join([f'risk_weight_{i}' for i in range(1, 25)])
                    cursor.execute(
                        f"""INSERT INTO almt_param_risk_weight (uuid, coa_cd, coa_name, weight, {risk_cols})
                            VALUES (%s, %s, %s, %s, {placeholders})""",
                        (str(uuid.uuid4()), coa_cd, coa_name, m1) + tuple(ws)
                    )
                    ins += 1
        conn.commit()
        return {"message": "导入成功", "inserted": ins, "updated": upd}
    finally:
        conn.close()


@router.get("/ftp-margin/export")
def export_ftp_margin():
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT coa_cd, coa_name, spread, curve_name, term FROM almt_param_ftp_spread ORDER BY coa_cd")
            rows = [list(r.values()) for r in cursor.fetchall()]
    finally:
        conn.close()
    return _export_excel(rows, ["账户册编码", "账户册名称", "FTP利差", "曲线名称", "期限"], "FTP利差", "ftp_margin.xlsx")


@router.post("/ftp-margin/import")
async def import_ftp_margin(file):
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    conn = get_db_conn()
    try:
        ins, upd = 0, 0
        with conn.cursor() as cursor:
            for row in rows:
                if not row or not row[0]:
                    continue
                vals = (list(row) + [None]*5)[:5]
                coa_cd, coa_name, spread, curve_name, term = vals
                cursor.execute("SELECT id FROM almt_param_ftp_spread WHERE coa_cd=%s", (str(coa_cd),))
                ex = cursor.fetchone()
                if ex:
                    cursor.execute(
                        "UPDATE almt_param_ftp_spread SET coa_name=%s, spread=%s, curve_name=%s, term=%s WHERE id=%s",
                        (str(coa_name or ''), float(spread or 0), str(curve_name or ''), str(term or ''), ex['id'])
                    )
                    upd += 1
                else:
                    cursor.execute(
                        "INSERT INTO almt_param_ftp_spread (uuid, coa_cd, coa_name, spread, curve_name, term) VALUES (%s, %s, %s, %s, %s, %s)",
                        (str(uuid.uuid4()), str(coa_cd), str(coa_name or ''), float(spread or 0), str(curve_name or ''), str(term or ''))
                    )
                    ins += 1
        conn.commit()
        return {"message": "导入成功", "inserted": ins, "updated": upd}
    finally:
        conn.close()


@router.get("/business-plan/export")
def export_business_plan():
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cols = ', '.join([f'plan_average{i}' for i in range(1, 25)])
            cursor.execute(f"SELECT coa_lvl, coa_cd, coa_name, {cols} FROM almt_param_business_plan ORDER BY coa_lvl")
            raw = cursor.fetchall()
            headers = ['层级', '账户册编码', '账户册名称'] + [f'M{i}日均增量' for i in range(1, 25)]
            rows = []
            for r in raw:
                row = [r.get('coa_lvl') or '', r.get('coa_cd') or '', r.get('coa_name') or '']
                for i in range(1, 25):
                    v = r.get(f'plan_average{i}')
                    row.append(float(v) if v is not None else None)
                rows.append(row)
    finally:
        conn.close()
    return _export_excel(rows, headers, "业务计划-日均增量", "business_plan_avg.xlsx")


@router.post("/business-plan/import")
async def import_business_plan(file):
    """导入业务计划，包含 M1~M24 日均增量（前3列：层级/编码/名称，后续24列：M1~M24日均增量）"""
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    conn = get_db_conn()
    try:
        ins, upd = 0, 0
        with conn.cursor() as cursor:
            for row in rows:
                if not row or len(row) < 2 or not row[1]:
                    continue
                vals = list(row) + [None] * (27 - len(row))
                coa_lvl, coa_cd, coa_name = (str(v) if v is not None else '' for v in vals[:3])
                avgs = []
                for i in range(3, 27):
                    try:
                        avgs.append(float(vals[i]) if vals[i] not in (None, '') else None)
                    except Exception:
                        avgs.append(None)
                cursor.execute("SELECT id FROM almt_param_business_plan WHERE coa_cd=%s", (coa_cd,))
                ex = cursor.fetchone()
                set_avg = ', '.join([f'plan_average{i}=%s' for i in range(1, 25)])
                if ex:
                    cursor.execute(
                        f"UPDATE almt_param_business_plan SET coa_lvl=%s, coa_name=%s, {set_avg} WHERE id=%s",
                        (coa_lvl, coa_name) + tuple(avgs) + (ex['id'],)
                    )
                    upd += 1
                else:
                    avg_cols = ', '.join([f'plan_average{i}' for i in range(1, 25)])
                    placeholders = ', '.join(['%s'] * 24)
                    cursor.execute(
                        f"""INSERT INTO almt_param_business_plan
                            (uuid, coa_lvl, coa_cd, coa_name, {avg_cols})
                            VALUES (%s, %s, %s, %s, {placeholders})""",
                        (str(uuid.uuid4()), coa_lvl, coa_cd, coa_name) + tuple(avgs)
                    )
                    ins += 1
        conn.commit()
        return {"message": "导入成功", "inserted": ins, "updated": upd}
    finally:
        conn.close()


@router.get("/rate-scenario/export")
def export_rate_scenario():
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT order_number, curve_name, curve_id, scenario_name, scenario_shift,
                          current_curve_value, m1_value, m2_value, m3_value, m4_value, m5_value, m6_value,
                          m7_value, m8_value, m9_value, m10_value, m11_value, m12_value,
                          m13_value, m14_value, m15_value, m16_value, m17_value, m18_value,
                          m19_value, m20_value, m21_value, m22_value, m23_value, m24_value
                FROM almt_param_rate_scenario ORDER BY scenario_name, order_number"""
            )
            rows = [list(r.values()) for r in cursor.fetchall()]
    finally:
        conn.close()
    headers = ["序号", "曲线名称", "曲线ID", "情景名称", "调整值", "当前值"]
    headers += ["M" + str(i) for i in range(1, 25)]
    return _export_excel(rows, headers, "利率情景", "rate_scenario.xlsx")


# ==================== 自定义定价策略（按账户册，每账户册一条） ====================


@router.get("/custom-strategy")
def list_custom_strategy():
    """获取定价策略列表（含 24 期策略值）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM almt_param_custom_strategy ORDER BY id LIMIT 100")
            return cursor.fetchall()
    finally:
        conn.close()


@router.get("/custom-strategy/tree")
def get_custom_strategy_tree():
    """按账户册树形展示定价策略（每账户册挂载 24 期策略值）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, coa_cd, coa_name, leaf_flag, parent_coa_cd FROM almt_coa_info")
            all_coa = cursor.fetchall()
            cols = ', '.join([f'strategy_M{i}' for i in range(1, 25)])
            cursor.execute(
                f"""SELECT coa_cd, coa_name, remark, {cols}, created_at, updated_at
                   FROM almt_param_custom_strategy"""
            )
            rows = cursor.fetchall()
            value_map = {r['coa_cd']: 1 for r in rows}
            strategy_map = {}
            for r in rows:
                strategy_map[r['coa_cd']] = {
                    'preview_values': [float(r[f'strategy_M{i}'] or 0) for i in range(1, 25)],
                    'last_update': str(r['updated_at']) if r['updated_at'] else None,
                    'remark': r['remark'] or ''
                }
        tree = build_tree_from_coa(all_coa, value_map)
        def attach_meta(nodes):
            for n in nodes:
                if n['coa_cd'] in strategy_map:
                    m = strategy_map[n['coa_cd']]
                    n.update(m)
                    n['has_strategy'] = True
                if n.get('children'):
                    attach_meta(n['children'])
        attach_meta(tree)
        return tree
    finally:
        conn.close()


@router.get("/custom-strategy/{coa_cd}")
def get_custom_strategy_by_coa(coa_cd: str):
    """获取某账户册的定价策略（coa_cd 唯一一条）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cols = ', '.join([f'strategy_M{i}' for i in range(1, 25)])
            cursor.execute(
                f"""SELECT id, uuid, coa_cd, coa_name, {cols}, remark
                   FROM almt_param_custom_strategy WHERE coa_cd=%s""",
                (coa_cd,)
            )
            return cursor.fetchone() or {}
    finally:
        conn.close()


@router.post("/custom-strategy/save")
def save_custom_strategy(item: dict):
    """按 coa_cd upsert 自定义定价策略（一个账户册一条）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            coa_cd = str(item.get('coa_cd', '')).strip()
            if not coa_cd:
                return {"message": "账户册编码不能为空", "success": False}
            coa_name = str(item.get('coa_name', '')).strip()
            remark = str(item.get('remark', '')).strip()
            m_vals = []
            for i in range(1, 25):
                v = item.get(f'strategy_M{i}')
                try:
                    m_vals.append(float(v) if v not in (None, '') else 0)
                except Exception:
                    m_vals.append(0)
            set_clause = ', '.join([f'strategy_M{i}=%s' for i in range(1, 25)])
            cursor.execute("SELECT id FROM almt_param_custom_strategy WHERE coa_cd=%s", (coa_cd,))
            ex = cursor.fetchone()
            if ex:
                cursor.execute(
                    f"""UPDATE almt_param_custom_strategy
                        SET coa_name=%s, {set_clause}, remark=%s
                        WHERE id=%s""",
                    (coa_name,) + tuple(m_vals) + (remark, ex['id'])
                )
                action = '更新'
            else:
                cols_m = ', '.join([f'strategy_M{i}' for i in range(1, 25)])
                placeholders = ', '.join(['%s'] * 24)
                cursor.execute(
                    f"""INSERT INTO almt_param_custom_strategy
                        (uuid, coa_cd, coa_name, {cols_m}, remark)
                        VALUES (%s, %s, %s, {placeholders}, %s)""",
                    (str(uuid.uuid4()), coa_cd, coa_name) + tuple(m_vals) + (remark,)
                )
                action = '创建'
        conn.commit()
        return {"message": f"{action}成功", "success": True}
    finally:
        conn.close()


@router.delete("/custom-strategy/{coa_cd}")
def delete_custom_strategy(coa_cd: str):
    """按 coa_cd 删除自定义定价策略"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM almt_param_custom_strategy WHERE coa_cd=%s", (coa_cd,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


@router.get("/custom-strategy/export")
def export_custom_strategy():
    """导出自定义定价策略：27 列（coa_cd/coa_name/remark + 24期策略值）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cols = ', '.join([f'strategy_M{i}' for i in range(1, 25)])
            cursor.execute(
                f"""SELECT coa_cd, coa_name, {cols}, remark
                   FROM almt_param_custom_strategy ORDER BY coa_cd"""
            )
            raw = cursor.fetchall()
            headers = ['账户册编码', '账户册名称']
            headers += [f'M{i}策略' for i in range(1, 25)]
            headers.append('备注')
            rows = []
            for r in raw:
                row = [r['coa_cd'] or '', r['coa_name'] or '']
                for i in range(1, 25):
                    v = r.get(f'strategy_M{i}')
                    row.append(float(v) if v is not None else 0)
                row.append(r.get('remark') or '')
                rows.append(row)
    finally:
        conn.close()
    return _export_excel(rows, headers, "自定义定价策略", "custom_strategy.xlsx")


@router.post("/custom-strategy/import")
async def import_custom_strategy(file):
    """导入自定义定价策略：2列基础信息 + 24列策略值 + 1列备注 = 27 列"""
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    conn = get_db_conn()
    try:
        ins, upd = 0, 0
        with conn.cursor() as cursor:
            for row in rows:
                if not row or not row[0]:
                    continue
                vals = list(row) + [None] * (27 - len(row))
                coa_cd = str(vals[0] or '').strip()
                if not coa_cd:
                    continue
                coa_name = str(vals[1] or '').strip()
                m_vals = []
                for i in range(2, 26):
                    try:
                        m_vals.append(float(vals[i]) if vals[i] not in (None, '') else 0)
                    except Exception:
                        m_vals.append(0)
                remark = str(vals[26] or '') if vals[26] is not None else ''

                set_clause = ', '.join([f'strategy_M{i}=%s' for i in range(1, 25)])
                cursor.execute("SELECT id FROM almt_param_custom_strategy WHERE coa_cd=%s", (coa_cd,))
                ex = cursor.fetchone()
                if ex:
                    cursor.execute(
                        f"""UPDATE almt_param_custom_strategy
                            SET coa_name=%s, {set_clause}, remark=%s
                            WHERE id=%s""",
                        (coa_name,) + tuple(m_vals) + (remark, ex['id'])
                    )
                    upd += 1
                else:
                    cols_m = ', '.join([f'strategy_M{i}' for i in range(1, 25)])
                    placeholders = ', '.join(['%s'] * 24)
                    cursor.execute(
                        f"""INSERT INTO almt_param_custom_strategy
                            (uuid, coa_cd, coa_name, {cols_m}, remark)
                            VALUES (%s, %s, %s, {placeholders}, %s)""",
                        (str(uuid.uuid4()), coa_cd, coa_name) + tuple(m_vals) + (remark,)
                    )
                    ins += 1
        conn.commit()
        return {"message": "导入成功", "inserted": ins, "updated": upd}
    finally:
        conn.close()