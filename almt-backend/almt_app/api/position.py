"""
存量数据API
"""
from fastapi import APIRouter
import pymysql
import uuid

router = APIRouter(prefix="/api/position", tags=["存量数据"])


# 统一使用 core.db_util，避免硬编码（云端与本地可通过 .env 切换账号）
from almt_app.core.db_util import get_db_conn


@router.get("")
def get_position_list(skip: int = 0, limit: int = 20, search: str = ''):
    """获取存量数据列表（支持分页和搜索）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            if search:
                sql = f"SELECT id, uuid, coa_lvl, coa_name, balance, average_balance, rate FROM almt_current_position WHERE coa_lvl LIKE %s OR coa_name LIKE %s LIMIT {limit} OFFSET {skip}"
                cursor.execute(sql, (f'%{search}%', f'%{search}%'))
            else:
                sql = f"SELECT id, uuid, coa_lvl, coa_name, balance, average_balance, rate FROM almt_current_position LIMIT {limit} OFFSET {skip}"
                cursor.execute(sql)
            return cursor.fetchall()
    finally:
        conn.close()


@router.get("/count")
def get_position_count(search: str = ''):
    """获取存量数据总数"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            if search:
                cursor.execute(
                    "SELECT COUNT(*) AS total FROM almt_current_position WHERE coa_lvl LIKE %s OR coa_name LIKE %s",
                    (f'%{search}%', f'%{search}%')
                )
            else:
                cursor.execute("SELECT COUNT(*) AS total FROM almt_current_position")
            row = cursor.fetchone()
            return {"total": row['total']}
    finally:
        conn.close()


@router.post("")
def create_position(item: dict):
    """创建存量数据"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            uuid_val = item.get('uuid') or str(uuid.uuid4())
            cursor.execute(
                """INSERT INTO almt_current_position (uuid, coa_lvl, coa_name, balance, average_balance, rate) 
                VALUES (%s, %s, %s, %s, %s, %s)""",
                (uuid_val, item.get('coa_lvl', ''), item.get('coa_name', ''),
                 item.get('balance', 0), item.get('average_balance', 0), item.get('rate', 0))
            )
        conn.commit()
        return {"message": "创建成功", "uuid": uuid_val}
    finally:
        conn.close()


@router.put("/{position_id}")
def update_position(position_id: int, item: dict):
    """更新存量数据"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """UPDATE almt_current_position 
                SET coa_lvl=%s, coa_name=%s, balance=%s, average_balance=%s, rate=%s 
                WHERE id=%s""",
                (item.get('coa_lvl', ''), item.get('coa_name', ''),
                 item.get('balance', 0), item.get('average_balance', 0), item.get('rate', 0),
                 position_id)
            )
        conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@router.delete("/{position_id}")
def delete_position(position_id: int):
    """删除存量数据"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM almt_current_position WHERE id=%s", (position_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


@router.get("/tree")
def get_position_tree():
    """获取按账户册树形结构组织的存量数据（联表）"""
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            # 获取所有账户册（包含parent_coa_cd用于构建父子关系）
            cursor.execute("SELECT id, coa_cd, coa_name, leaf_flag, parent_coa_cd FROM almt_coa_info")
            all_coa = cursor.fetchall()

            # 获取所有存量数据（按 coa_lvl 聚合）
            cursor.execute("""
                SELECT coa_lvl, SUM(balance) as balance, SUM(average_balance) as average_balance,
                       AVG(rate) as rate, COUNT(*) as cnt
                FROM almt_current_position
                GROUP BY coa_lvl
            """)
            position_map = {}
            for row in cursor.fetchall():
                position_map[row['coa_lvl']] = row

        # 构建账户册节点
        coa_map = {}
        roots = []

        for coa in all_coa:
            node = {
                "id": coa['id'],
                "coa_cd": coa['coa_cd'],
                "coa_name": coa['coa_name'],
                "leaf_flag": coa['leaf_flag'],
                "balance": None,
                "average_balance": None,
                "rate": None,
                "has_data": False,
                "children": []
            }
            coa_map[coa['coa_cd']] = node

            # 精确匹配存量数据
            if coa['coa_cd'] in position_map:
                pos = position_map[coa['coa_cd']]
                node['balance'] = float(pos['balance'] or 0)
                node['average_balance'] = float(pos['average_balance'] or 0)
                node['rate'] = float(pos['rate'] or 0)
                node['has_data'] = True

        # 使用 parent_coa_cd 构建父子关系
        for coa in all_coa:
            coa_cd = coa['coa_cd']
            node = coa_map[coa_cd]
            parent_cd = coa['parent_coa_cd']

            if not parent_cd or parent_cd not in coa_map:
                # 父节点不存在，作为根节点
                roots.append(node)
            else:
                coa_map[parent_cd]['children'].append(node)

        # 自底向上汇总
        def calc_sum(node):
            for child in node['children']:
                calc_sum(child)
            if not node['children']:
                return
            total_bal = 0
            total_avg = 0
            any_data = False
            for child in node['children']:
                if child['balance'] is not None:
                    total_bal += child['balance']
                    any_data = True
                if child['average_balance'] is not None:
                    total_avg += child['average_balance']
            if any_data:
                node['balance'] = total_bal
                node['average_balance'] = total_avg
                node['has_data'] = True

        for root in roots:
            calc_sum(root)

        return roots
    finally:
        conn.close()


# ============ 导入导出 ============
@router.get("/export")
def export_position():
    """导出存量数据为Excel"""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io

    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT coa_lvl, coa_name, balance, average_balance, rate FROM almt_current_position ORDER BY coa_lvl")
            rows = cursor.fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "存量数据"
    ws.append(["层级编码", "账户册名称", "余额", "日均余额", "利率"])
    for r in rows:
        ws.append([r['coa_lvl'], r['coa_name'], r['balance'] or 0, r['average_balance'] or 0, r['rate'] or 0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=position.xlsx"}
    )


@router.post("/import")
async def import_position(file):
    """从Excel导入存量数据"""
    from openpyxl import load_workbook
    import io

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    conn = get_db_conn()
    try:
        inserted = 0
        updated = 0
        with conn.cursor() as cursor:
            for row in rows:
                if not row or not row[0]:
                    continue
                coa_lvl, coa_name, balance, avg_balance, rate = (list(row) + [None]*5)[:5]
                cursor.execute("SELECT id FROM almt_current_position WHERE coa_lvl=%s", (str(coa_lvl),))
                existing = cursor.fetchone()
                if existing:
                    cursor.execute(
                        """UPDATE almt_current_position SET coa_name=%s, balance=%s, average_balance=%s, rate=%s WHERE id=%s""",
                        (str(coa_name or ''), float(balance or 0), float(avg_balance or 0), float(rate or 0), existing['id'])
                    )
                    updated += 1
                else:
                    cursor.execute(
                        """INSERT INTO almt_current_position (uuid, coa_lvl, coa_name, balance, average_balance, rate) 
                        VALUES (%s, %s, %s, %s, %s, %s)""",
                        (str(uuid.uuid4()), str(coa_lvl), str(coa_name or ''),
                         float(balance or 0), float(avg_balance or 0), float(rate or 0))
                    )
                    inserted += 1
        conn.commit()
        return {"message": "导入成功", "inserted": inserted, "updated": updated}
    finally:
        conn.close()
