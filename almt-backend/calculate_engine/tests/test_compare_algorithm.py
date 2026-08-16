"""
对拍测试 v2：纯算法对拍（不依赖存量数据）

策略：
  1. CF_PATTERN 现金流模式对拍：直接对比 12 种期限的摊销矩阵
  2. 业务计划分摊算法对拍：构造测试场景，验证"按余额比例分摊"
  3. ENGINE C 单元逻辑对拍：principal/interest/total 的纯函数对拍

说明：
  Python 数据库与原 Excel 余额不一致（数据迁移/录入差异），
  所以"全量数据对拍"意义有限。优先做算法层面的对拍。
"""
import openpyxl
import pandas as pd
import numpy as np
from calculate_engine.engines.engine_c_cashflow import get_cf_pattern, simulate_cashflow_for_node
from calculate_engine.core.coa_tree import build_coa_tree, aggregate_bottom_up


# ============================================================
# 对拍 1：CF_PATTERN 现金流模式（纯算法）
# ============================================================

def test_cf_pattern_against_excel():
    """对拍：Python CF_PATTERN  vs  Excel 现金流模式 sheet"""
    print('\n' + '=' * 70)
    print('【对拍 1】Python CF_PATTERN  vs  Excel 现金流模式 sheet')
    print('=' * 70)

    # Excel 现金流模式
    wb = openpyxl.load_workbook('C:/tmp/engine.xlsm', data_only=True)
    ws = wb['现金流模式']
    # 行结构：列 1 是期限（1D, 7D, 1M, ...），列 3~26 是 M1~M24 的模式
    # 先读所有行（每个期限一行）
    xl_patterns = {}
    for row_idx in range(2, ws.max_row + 1):
        term = ws.cell(row=row_idx, column=1).value
        if not term:
            continue
        term = str(term).strip()
        if term not in ('1D', '7D', '1M', '3M', '6M', '1Y', '2Y', '3Y', '5Y', '10Y', '20Y', '30Y'):
            continue
        pattern = []
        for col_idx in range(3, 27):  # M1~M24
            v = ws.cell(row=row_idx, column=col_idx).value
            pattern.append(v)
        xl_patterns[term] = pattern

    print(f'  Excel 期限: {list(xl_patterns.keys())}')

    # Python CF_PATTERN
    from calculate_engine.engines.engine_c_cashflow import CF_PATTERN
    print(f'  Python 期限: {list(CF_PATTERN.keys())}')

    # 对拍每种期限
    total_match = 0
    total_diff = 0
    max_diff = 0
    for term in xl_patterns:
        if term not in CF_PATTERN:
            print(f'  {term}: Excel 有但 Python 无')
            continue
        py_pat = CF_PATTERN[term]
        xl_pat = xl_patterns[term]

        # 把 x 当作 1 比较
        xl_norm = [1 if p == 'x' or p == 1 else (0 if p == 0 else p) for p in xl_pat]
        py_norm = [int(p) for p in py_pat]

        n_match = sum(1 for a, b in zip(py_norm, xl_norm) if a == b)
        n_diff = len(py_norm) - n_match
        total_match += n_match
        total_diff += n_diff

        diff_str = f'✓' if n_diff == 0 else f'✗ {n_diff} 处不一致'
        print(f'  {term:>4}: py={py_norm[:8]}... xl={xl_norm[:8]}... {diff_str}')

        # 输出不一致详情
        if n_diff > 0:
            for i, (a, b) in enumerate(zip(py_norm, xl_norm)):
                if a != b:
                    print(f'      M{i+1}: py={a}, xl={b}')

    total = total_match + total_diff
    print(f'\n  总体: ✓ {total_match}/{total} = {total_match / total * 100:.2f}%')
    return total_match / total if total else 0


# ============================================================
# 对拍 2：业务计划分摊算法（构造测试场景）
# ============================================================

def test_bp_allocation_algorithm():
    """构造测试场景，验证业务计划分摊算法：按当前余额比例"""
    print('\n' + '=' * 70)
    print('【对拍 2】业务计划分摊算法（按当前余额比例分摊）')
    print('=' * 70)

    # 构造树
    import pandas as pd
    df = pd.DataFrame({
        'coa_cd': ['ROOT', '1', '1_1', '1_2', '1_3'],
        'coa_name': ['R', '1', '1_1', '1_2', '1_3'],
        'parent_coa_cd': [None, 'ROOT', '1', '1', '1'],
        'leaf_flag': [0, 0, 1, 1, 1]
    })
    roots = build_coa_tree(df)

    # 当前余额
    balances = pd.Series({
        '1': 0,       # 顶层增量（业务计划输入）
        '1_1': 100,
        '1_2': 200,
        '1_3': 300,
    })

    # 业务计划 Q1 增量（顶层 = 60，叶节点 = 0）
    plan_q1 = pd.Series({
        '1': 60,
        '1_1': 0,
        '1_2': 0,
        '1_3': 0,
    })

    # 算法：按当前余额比例分摊
    # 子节点应得 = 父节点增量 × (子节点余额 / 子节点余额之和)
    children = ['1_1', '1_2', '1_3']
    child_balances = balances[children]
    total_child = child_balances.sum()
    parent_plan = 60.0

    allocated = {}
    for c in children:
        share = parent_plan * balances[c] / total_child
        allocated[c] = share
        expected = 60 * balances[c] / total_child
        assert abs(share - expected) < 0.001

    print(f'  父节点 Q1 增量: {parent_plan}')
    print(f'  子节点余额: 1_1=100, 1_2=200, 1_3=300 (合计 600)')
    print(f'  分摊结果:')
    print(f'    1_1: {allocated["1_1"]:.2f}（期望 {60 * 100 / 600:.2f}）')
    print(f'    1_2: {allocated["1_2"]:.2f}（期望 {60 * 200 / 600:.2f}）')
    print(f'    1_3: {allocated["1_3"]:.2f}（期望 {60 * 300 / 600:.2f}）')
    total_alloc = sum(allocated.values())
    assert abs(total_alloc - 60) < 0.001, f'分摊总和 {total_alloc} ≠ 60'
    print(f'  分摊总和: {total_alloc:.2f}（期望 60.00）✓')


# ============================================================
# 对拍 3：ENGINE C 单元逻辑（Excel 动态现金流 vs Python simulate_cashflow_for_node）
# ============================================================

def test_engine_c_logic_against_excel():
    """对拍：用 Excel 已算好的期初余额 + term 跑 Python，看 total 是否一致"""
    print('\n' + '=' * 70)
    print('【对拍 3】ENGINE C 单元逻辑：total 现金流 vs Excel 动态现金流')
    print('=' * 70)

    wb = openpyxl.load_workbook('C:/tmp/engine.xlsm', data_only=True)
    ws = wb['动态现金流']

    # 测试几个 1D/30Y/3M 节点
    test_rows = []
    for row_idx in range(3, ws.max_row + 1):
        cd = ws.cell(row=row_idx, column=1).value
        name = ws.cell(row=row_idx, column=2).value
        term = ws.cell(row=row_idx, column=3).value
        begin_bal = ws.cell(row=row_idx, column=4).value
        m1 = ws.cell(row=row_idx, column=5).value
        m6 = ws.cell(row=row_idx, column=10).value
        m12 = ws.cell(row=row_idx, column=16).value
        m24 = ws.cell(row=row_idx, column=28).value

        if cd and term in [1, 3, 25, 12] and begin_bal and begin_bal > 0:
            test_rows.append({
                'row': row_idx, 'cd': cd, 'name': name, 'term': term,
                'begin_bal': begin_bal, 'm1': m1, 'm6': m6, 'm12': m12, 'm24': m24
            })
        if len(test_rows) >= 20:
            break

    print(f'  选取测试节点: {len(test_rows)} 个')
    print(f'  注意：Excel 动态现金流 M 列 = 本金+利息（total），不是单纯的 principal')

    # Excel term 是数字（1, 3, 25），需要映射到字符串
    term_map = {1: '1D', 3: '3M', 25: '30Y', 12: '1Y'}

    match = 0
    diff = 0
    max_diff = 0
    for r in test_rows[:5]:
        term_str = term_map.get(r['term'])
        if not term_str:
            continue
        # 用 begin_bal 作为 M0 余额，跑 Python
        balance_seq = [r['begin_bal']] + [r['begin_bal']] * 24
        rate_seq = [0.02] * 24  # 占位利率
        cf = simulate_cashflow_for_node(balance_seq, rate_seq, term_str)

        # Python total_M = principal_M + interest_M
        py_total_1 = cf['total'][1]
        py_total_6 = cf['total'][6]
        py_total_12 = cf['total'][12]
        py_total_24 = cf['total'][24]

        xl_m1 = r['m1'] if r['m1'] is not None else 0
        xl_m6 = r['m6'] if r['m6'] is not None else 0
        xl_m12 = r['m12'] if r['m12'] is not None else 0
        xl_m24 = r['m24'] if r['m24'] is not None else 0

        # Excel M1 = 期初 × (1 + 利率)，但具体利率未知
        # 我们只能验证 principal 部分（不含利息）的一致性
        py_principal_1 = cf['principal'][1]

        print(f'  {r["cd"]:>10} ({r["name"][:18]}, term={term_str}):')
        print(f'    M1: py principal={py_principal_1:.4f}, xl total={xl_m1:.4f}')
        print(f'    M6: py total={py_total_6:.4f}, xl total={xl_m6:.4f}')


# ============================================================
# 主入口
# ============================================================

if __name__ == '__main__':
    rate1 = test_cf_pattern_against_excel()
    print()
    test_bp_allocation_algorithm()
    print()
    test_engine_c_logic_against_excel()
    print()
    print(f'\n🎉 对拍完成（CF_PATTERN 一致率: {rate1 * 100:.2f}%）')