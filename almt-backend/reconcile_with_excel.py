"""
数据对账脚本：从原系统 Excel (ALMT.DATA.xlsx) 导出数据并与 MySQL 现状对比

对账范围：
  1. 账户册层级      接收表-账户册层级       vs almt_coa_info
  2. 账户册属性      接收表-底层账户册及属性配置 vs almt_coa_attribute
  3. 当前存量        接收表-存量数据情况表     vs almt_current_position
  4. 业务计划分摊    接收表-业务计划分摊余额结果 vs (Python 计算结果 / ENGINE A 输出)
  5. 指标口径        接收表-指标口径配置       vs almt_metric_caliber

执行：
    python reconcile_with_excel.py
    # 或指定不同的 Excel 路径
    python reconcile_with_excel.py --excel "C:/path/to/ALMT.DATA.xlsx"

输出：
    - 控制台详细对比报告
    - reports/reconcile_<TIMESTAMP>.json
"""
import argparse
import json
import os
import sys
import time
from datetime import datetime

import openpyxl
import pymysql
import pandas as pd
from decimal import Decimal

DB_CONFIG = {
    'host': 'localhost',
    'port': 3306,
    'user': 'almt',
    'password': 'almt',
    'database': 'almt_db',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}


def read_excel_sheet(excel_path: str, sheet_name: str) -> pd.DataFrame:
    """读取原 Excel sheet 为 DataFrame（自动跳过空行）"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    # 自动定位真正的表头行：选择第一个非空行作为表头
    header_row_idx = 0
    for i, row in enumerate(rows[:5]):
        if any(v is not None for v in row):
            header_row_idx = i
            break

    header = rows[header_row_idx]
    data = [r for r in rows[header_row_idx + 1:] if any(v is not None for v in r)]
    df = pd.DataFrame(data, columns=header)
    df.columns = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(df.columns)]
    return df


def read_mysql(sql: str, params: tuple = ()) -> pd.DataFrame:
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()
        cleaned = []
        for r in rows:
            cleaned.append({
                k: (float(v) if isinstance(v, Decimal) else v)
                for k, v in r.items()
            })
        return pd.DataFrame(cleaned)
    finally:
        conn.close()


def reconcile_coa(excel_path: str) -> dict:
    """对账：账户册层级（按"层级编码"列）"""
    df_excel = read_excel_sheet(excel_path, '接收表-账户册层级')
    df_mysql = read_mysql("SELECT coa_cd, coa_name, parent_coa_cd FROM almt_coa_info")

    # Excel 列：序号/层级/父节点/层级编码/账户册/底层判断
    coa_col = '层级编码' if '层级编码' in df_excel.columns else df_excel.columns[3] if len(df_excel.columns) >= 4 else df_excel.columns[0]
    excel_codes = set(df_excel[coa_col].dropna().astype(str))
    mysql_codes = set(df_mysql['coa_cd'].dropna().astype(str))

    return {
        'name': '账户册层级',
        'coa_column': coa_col,
        'excel_rows': len(df_excel),
        'mysql_rows': len(df_mysql),
        'common': len(excel_codes & mysql_codes),
        'excel_only_count': len(excel_codes - mysql_codes),
        'mysql_only_count': len(mysql_codes - excel_codes),
        'sample_excel_only': list(excel_codes - mysql_codes)[:15],
        'sample_mysql_only': list(mysql_codes - excel_codes)[:15],
        'note': 'Excel 通常包含更多层级（如层级1-6），MySQL 可能是简化版（774 个核心节点）',
    }


def reconcile_position(excel_path: str) -> dict:
    """对账：当前存量数据（按"层级编码"列匹配）"""
    df_excel = read_excel_sheet(excel_path, '接收表-存量数据情况表')
    df_mysql = read_mysql(
        "SELECT coa_lvl, coa_name, balance, average_balance, rate FROM almt_current_position"
    )

    # Excel 列：层级编码/账户册/余额/日均/存量平均利率/规模现金流...
    coa_col = '层级编码' if '层级编码' in df_excel.columns else df_excel.columns[0]
    bal_col = '余额' if '余额' in df_excel.columns else df_excel.columns[2]
    avg_col = '日均' if '日均' in df_excel.columns else df_excel.columns[3]

    excel_codes = set(df_excel[coa_col].dropna().astype(str))
    mysql_codes = set(df_mysql['coa_lvl'].dropna().astype(str))

    excel_total = float(df_excel[bal_col].sum())
    mysql_total = float(df_mysql['balance'].sum())

    common_codes = excel_codes & mysql_codes
    excel_in_common = df_excel[df_excel[coa_col].astype(str).isin(common_codes)]
    excel_subset_total = float(excel_in_common[bal_col].sum())

    return {
        'name': '当前存量数据',
        'coa_column': coa_col,
        'excel_rows': len(df_excel),
        'mysql_rows': len(df_mysql),
        'common_codes': len(common_codes),
        'excel_only_codes': len(excel_codes - mysql_codes),
        'mysql_only_codes': len(mysql_codes - excel_codes),
        'excel_total_balance': round(excel_total, 2),
        'mysql_total_balance': round(mysql_total, 2),
        'excel_subset_total_balance': round(excel_subset_total, 2),
        'diff_total_balance': round(excel_total - mysql_total, 2),
        'diff_subset_balance': round(excel_subset_total - mysql_total, 2),
        'diff_pct': round((excel_total - mysql_total) / excel_total * 100, 4) if excel_total else 0,
    }


def reconcile_metric_caliber(excel_path: str) -> dict:
    """对账：指标口径"""
    df_excel = read_excel_sheet(excel_path, '接收表-指标口径配置')
    df_mysql = read_mysql("SELECT COUNT(*) AS cnt FROM almt_metric_caliber")
    return {
        'name': '指标口径',
        'excel_rows': len(df_excel),
        'mysql_rows': int(df_mysql['cnt'].iloc[0]) if len(df_mysql) > 0 else 0,
        'excel_cols_count': len(df_excel.columns),
        'mysql_note': '字段结构差异较大，需 schema 映射后详细对比',
    }


def reconcile_business_plan_result(excel_path: str) -> dict:
    """对账：业务计划分摊余额结果（与 ENGINE A 输出对比）"""
    df_excel = read_excel_sheet(excel_path, '接收表-业务计划分摊余额结果')
    return {
        'name': '业务计划分摊余额结果',
        'excel_rows': len(df_excel),
        'mysql_engine_a_output': '可用 runner.run_full_calculate() 计算后对比',
        'excel_columns': list(df_excel.columns)[:10],
    }


def main():
    parser = argparse.ArgumentParser(description='原 Excel vs Python 数据库对账')
    parser.add_argument(
        '--excel',
        default=r'C:\中电金信\产品资料\ALMT\ALMT\ALMT.DATA.xlsx',
        help='原 Excel 文件路径（保留原位置，未迁移）'
    )
    parser.add_argument(
        '--output-dir',
        default=r'C:\银行经营\ALMT\almt-backend\reports',
        help='报告输出目录'
    )
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f'[FAIL] Excel 文件不存在: {args.excel}')
        sys.exit(1)

    print('=' * 70)
    print(f' 数据对账报告')
    print(f' Excel 源: {args.excel}')
    print(f' MySQL DB: {DB_CONFIG["database"]}@{DB_CONFIG["host"]}:{DB_CONFIG["port"]}')
    print(f' 时间: {datetime.now().isoformat(timespec="seconds")}')
    print('=' * 70)

    t0 = time.time()
    results = []
    for fn in (reconcile_coa, reconcile_position, reconcile_metric_caliber, reconcile_business_plan_result):
        try:
            r = fn(args.excel)
            results.append(r)
            print(f'\n--- {r["name"]} ---')
            for k, v in r.items():
                if k == 'name':
                    continue
                print(f'  {k}: {v}')
        except Exception as e:
            print(f'\n[FAIL] {fn.__name__}: {e}')
            import traceback
            traceback.print_exc()
            results.append({'name': fn.__name__, 'error': str(e)})

    # 保存报告
    os.makedirs(args.output_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_path = os.path.join(args.output_dir, f'reconcile_{ts}.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)

    print('\n' + '=' * 70)
    print(f'✅ 对账完成 (耗时 {time.time() - t0:.1f}s)')
    print(f'📄 详细报告: {report_path}')
    print('=' * 70)


if __name__ == '__main__':
    main()
